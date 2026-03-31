from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


SHEET_KEYS = {
    "Required Core Courses": "requiredCoreCourses",
    "CISE Technical Electives": "ciseTechnicalElectives",
    "Approved Tech Electives": "approvedTechElectives",
}


def clean_group_label(value: str) -> str:
    return value.replace(chr(0x2500), "").strip()


def parse_courses(workbook_path: Path) -> dict[str, object]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    catalog: dict[str, object] = {}

    for sheet_name, key in SHEET_KEYS.items():
        worksheet = workbook[sheet_name]
        current_group: str | None = None
        courses: list[dict[str, object]] = []

        for index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            if index == 1:
                continue

            code, name, difficulty, notes = row[:4]
            if code is None and name is None and difficulty is None and notes is None:
                continue

            if name is None and difficulty is None and notes is None:
                current_group = clean_group_label(str(code))
                continue

            display_code = str(code).strip()
            courses.append(
                {
                    "code": display_code.replace(" ", ""),
                    "displayCode": display_code,
                    "name": name,
                    "difficulty": difficulty,
                    "notes": notes,
                    "group": current_group,
                }
            )

        catalog[key] = courses

    legend_rows = workbook["Legend"].iter_rows(min_row=3, max_row=5, values_only=True)
    legend = [
        {"difficulty": label, "description": description}
        for label, description in legend_rows
        if label
    ]

    note = next(
        (
            description
            for label, description in workbook["Legend"].iter_rows(values_only=True)
            if label == "Note:"
        ),
        None,
    )

    return {"legend": legend, "note": note, **catalog}


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Convert the UF curriculum workbook into repo-local JSON."
    )
    parser.add_argument("workbook", type=Path, help="Path to the .xlsx workbook")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("data/course-catalog.json"),
        help="Destination JSON path",
    )
    args = parser.parse_args()

    payload = parse_courses(args.workbook)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
